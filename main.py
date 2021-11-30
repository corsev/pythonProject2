import mysql.connector as conector
import bs4
from bs4 import BeautifulSoup
import requests
import openpyxl
from tkinter import *
from tkinter import ttk


def cargar_todas_paginas():
    soups = []
    num_pag = 0
    for a in range(71):
        url_todas = "https://www.ufcespanol.com/athletes/all?filters%5B0%5D=status%3A23&gender=All&search=&page=" + str(
            num_pag) + '"'
        pagina = requests.get(url_todas)
        soup = BeautifulSoup(pagina.text, "html.parser")
        soups.append(soup)
        num_pag += 1

    return soups


def cargar_datos():
    soups = cargar_todas_paginas()
    lista_datos_luchadores = list()
    lista_cara = list()
    lista_nombre = list()
    lista_apodo = list()
    lista_peso = list()
    lista_ratio = list()
    for soup in soups:

        datos_obtener_frontal = soup.findAll("div", {"class": "c-listing-athlete-flipcard__front"})

        # Creamos la plantilla del diccionario con los datos del luchador a buscar

        luchador = {
            "cara": "",
            "nombre": "",
            "apodo": "",
            "peso": "",
            "ratio": "",
            # "cuerpo" : "",
        }

        # Buscamos los elementos de los luchadores de la parte delantera y lo guardamos en el diccionario

        for elemento in datos_obtener_frontal:
            cara = "\"Sin imagen\""
            if elemento.find("div", {"class": "layout__region layout__region--content"}) is not None:
                cara = elemento.find("div", {"class": "layout__region layout__region--content"}).find("img").attrs[
                    "src"]

            nombre = str(elemento.find("div", {"class": "c-listing-athlete__text"}).find("span",
                                                                                         "c-listing-athlete__name").text.replace(
                '\n', ' ').lstrip().rstrip())
            apodo = "\"Sin apodo\""
            if elemento.find("div", {
                "class": "field field--name-nickname field--type-string field--label-hidden"}) is not None:
                apodo = elemento.find("div", {
                    "class": "field field--name-nickname field--type-string field--label-hidden"}).find("div", {
                    "class": "field__item"}).text
            peso = elemento.find("div", {
                "class": "field field--name-stats-weight-class field--type-entity-reference field--label-hidden field__items"}).find(
                "div", {"class": "field__item"}).text
            ratio = elemento.find("div", {"class": "c-listing-athlete__text"}).find("span", {
                "class": "c-listing-athlete__record"}).text

            # Guardamos los datos extraidos y los metemos en diccionarios y listas

            luchadores_datos = luchador.copy()
            luchadores_datos["cara"] = cara
            luchadores_datos["nombre"] = nombre
            luchadores_datos["apodo"] = apodo
            luchadores_datos["peso"] = peso
            luchadores_datos["ratio"] = ratio
            lista_cara.append(cara)
            lista_nombre.append(nombre)
            lista_apodo.append(apodo)
            lista_peso.append(peso)
            lista_ratio.append(ratio)

            # Guardamos el diccionario en una lista para guardar a todos los luchadores
            lista_datos_luchadores.append(luchadores_datos)

        # Creamos un excel con todos los datos de los jugadores

        wb = openpyxl.Workbook()
        ruta = 'Luchadores UFC.xlsx'
        hoja = wb.active
        hoja.title = "Datos Luchadores"
        hoja.append(('Nombre', 'Apodo', 'Peso', 'Ratio', 'Cara'))

        fila = 2

        col_cara = 5
        col_nombre = 1
        col_apodo = 2
        col_peso = 3
        col_ratio = 4

        for caras, nombres, apodos, pesos, ratios in zip(lista_cara, lista_nombre, lista_apodo, lista_peso,
                                                         lista_ratio):
            hoja.cell(column=col_cara, row=fila, value=caras)
            hoja.cell(column=col_nombre, row=fila, value=nombres)
            hoja.cell(column=col_apodo, row=fila, value=apodos)
            hoja.cell(column=col_peso, row=fila, value=pesos)
            hoja.cell(column=col_ratio, row=fila, value=ratios)
            fila += 1

            # wb.save(filename=ruta)

    return lista_datos_luchadores


def conectar_bbdd():
    conexion = conector.connect(user='root',
                                password='1234',
                                host='localhost',
                                database='luchadores_ufc',
                                autocommit=True)

    return conexion


def insertar_datos():
    conexion = conectar_bbdd()
    cursor = conexion.cursor()
    datos_luchadores = cargar_datos()
    ql_insertar_datos = "INSERT INTO luchadores (nombre,apodo,peso,ratio,cara) VALUES (%s, %s, %s, %s, %s)"
    for luchador in datos_luchadores:
        values = [luchador["nombre"], luchador["apodo"], luchador["peso"], luchador["ratio"], luchador["cara"]]
        cursor.execute(ql_insertar_datos, values)
    print("Se han metido todos los datos")
    cursor.close()


def eliminar_datos():
    conexion = conectar_bbdd()
    cursor = conexion.cursor()
    eliminar = "DELETE FROM luchadores"
    cursor.execute(eliminar)
    print("Se han eliminado todos los datos")
    cursor.close()


def consultar_datos():
    lista_datos_luchadores = list()
    conexion = conectar_bbdd()
    cursor = conexion.cursor()
    consultar_todo = "SELECT * FROM luchadores"
    cursor.execute(consultar_todo)
    datos = cursor.fetchall()
    for ids, nombres, apodos, pesos, ratios, caras in datos:
        luchador = {"id": ids, "cara": caras, "nombre": nombres, "apodo": apodos, "peso": pesos, "ratio": ratios}
        lista_datos_luchadores.append(luchador)

    print("Se han recuperado todos los datos")
    cursor.close()

    return lista_datos_luchadores


def ventana_mostrar():
    ventana2 = Tk()
    menu_ventana = Menu(ventana2)
    ventana2.title("Ventana para mostrar los datos")
    ventana2.config(menu=menu_ventana)
    ventana2.geometry("1400x200")
    ventana2.resizable(True, True)
    luchadores = consultar_datos()
    tv = ttk.Treeview(ventana2)

    tv['columns'] = ("id", "nombre", "apodo", "peso", "ratio", "cara")

    tv.column("#0", width=0, anchor=CENTER)
    tv.column("id", width=50, anchor=CENTER)
    tv.column("nombre", width=150, anchor=CENTER)
    tv.column("apodo", width=150, anchor=CENTER)
    tv.column("peso", width=150, anchor=CENTER)
    tv.column("ratio", width=150, anchor=CENTER)
    tv.column("cara", width=800, anchor=CENTER)

    tv.heading("#0", text="", anchor=CENTER)
    tv.heading("id", text="Id", anchor=CENTER)
    tv.heading("nombre", text="Nombre", anchor=CENTER)
    tv.heading("apodo", text="Apodo", anchor=CENTER)
    tv.heading("peso", text="Peso", anchor=CENTER)
    tv.heading("ratio", text="Ratio", anchor=CENTER)
    tv.heading("cara", text="Cara", anchor=CENTER)

    siguiente = 1

    for luchadore in luchadores:
        tv.insert(parent="", index=siguiente, values=(luchadore["id"], luchadore["nombre"], luchadore["apodo"],
                                                      luchadore["peso"], luchadore["ratio"], luchadore["cara"]))
        siguiente += 1

    barra1 = Scrollbar(ventana2, command=tv.yview)
    barra1.pack(side=RIGHT, fill=Y)
    tv.config(yscrollcommand=barra1.set)

    tv.pack()
    ventana2.mainloop()


def insertar_luchador():
    ventana3 = Tk()
    ventana3.geometry("500x500")
    ventana3.title("Guardar Luchador")
    inserta = Label(ventana3, text="Guarda Aquí Luchadores", bg="grey", fg="black", width="500", height="3")

    nombre_texto = Label(ventana3, text="Nombre")
    apodo_texto = Label(ventana3, text="Apodo")
    peso_texto = Label(ventana3, text="Peso")
    ratio_texto = Label(ventana3, text="Ratio")
    cara_texto = Label(ventana3, text="Cara")

    nombre_texto.place(x=15, y=70)
    apodo_texto.place(x=15, y=140)
    peso_texto.place(x=15, y=210)
    ratio_texto.place(x=15, y=280)
    cara_texto.place(x=15, y=350)

    nombre = StringVar()
    apodo = StringVar()
    peso = StringVar()
    ratio = StringVar()
    cara = StringVar()

    nombre_entrada = Entry(ventana3, textvariable=nombre, width="30")
    apodo_entrada = Entry(ventana3, textvariable=apodo, width="30")
    peso_entrada = Entry(ventana3, textvariable=peso, width="30")
    ratio_entrada = Entry(ventana3, textvariable=ratio, width="30")
    cara_entrada = Entry(ventana3, textvariable=cara, width="30")

    nombre_entrada.place(x=15, y=100)
    apodo_entrada.place(x=15, y=170)
    peso_entrada.place(x=15, y=240)
    ratio_entrada.place(x=15, y=310)
    cara_entrada.place(x=15, y=380)

    guardar = Button(ventana3, text="Guardar", width=30, height=2, bg="grey",
                     command=lambda: insertar_luchador_tkinter([nombre_entrada.get(),
                                                                apodo_entrada.get(),
                                                                peso_entrada.get(),
                                                                ratio_entrada.get(),
                                                                cara_entrada.get()]))
    guardar.place(x=15, y=430)

    inserta.pack()
    ventana3.mainloop()

    return


def insertar_luchador_tkinter(nombre_entrada, apodo_entrada, peso_entrada, ratio_entrada, cara_entrada):
    return


def aplicacion_luchadores():
    root = Tk()
    menu_principal = Menu(root)
    root.title("Luchadores UFC")
    root.config(menu=menu_principal)
    root.geometry("800x0")
    root.resizable(True, True)

    barra_menu = Menu(menu_principal, tearoff=0)
    barra_menu1 = Menu(menu_principal, tearoff=0)
    barra_menu2 = Menu(menu_principal, tearoff=0)
    barra_menu3 = Menu(menu_principal, tearoff=0)

    menu_principal.add_cascade(label="Cargar", menu=barra_menu)
    barra_menu.add_command(label="Luchadores", command=lambda: insertar_datos())

    menu_principal.add_cascade(label="Eliminar", menu=barra_menu1)
    barra_menu1.add_command(label="Luchadores", command=lambda: eliminar_datos())

    menu_principal.add_cascade(label="Mostrar", menu=barra_menu2)
    barra_menu2.add_command(label="Luchadores", command=lambda: ventana_mostrar())

    menu_principal.add_cascade(label="Introducir", menu=barra_menu3)
    barra_menu3.add_command(label="Luchadores", command=lambda: insertar_luchador())

    root.mainloop()


aplicacion_luchadores()
